from __future__ import annotations

import io
import os
import stat
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))

from tools.testkit import assert_launcher_help, load_launcher


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "expense_dock"


def load_module(monkeypatch, runtime_home: Path):
    monkeypatch.setenv("EXPENSE_DOCK_HOME", str(runtime_home))
    return load_launcher(SCRIPT_PATH)


def test_help_does_not_create_runtime_state(tmp_path):
    runtime_home = tmp_path / "runtime_home"
    assert_launcher_help(
        SCRIPT_PATH,
        env_overrides={"EXPENSE_DOCK_HOME": str(runtime_home)},
    )

    assert not runtime_home.exists()


def test_runtime_config_is_seeded(monkeypatch, tmp_path):
    module = load_module(monkeypatch, tmp_path / "runtime_home")

    module.ensure_runtime_dirs()
    module.ensure_runtime_config()

    assert Path(module.CONFIG_FILE).exists()
    assert stat.S_IMODE(Path(module.DATA_DIR).stat().st_mode) == 0o700
    assert stat.S_IMODE(Path(module.PENDING_DIR).stat().st_mode) == 0o700
    assert stat.S_IMODE(Path(module.CONFIG_FILE).stat().st_mode) == 0o600


def test_existing_runtime_permissions_are_hardened(monkeypatch, tmp_path):
    runtime_home = tmp_path / "runtime_home"
    pending = runtime_home / "pending"
    pending.mkdir(parents=True)
    config = runtime_home / "config.json"
    config.write_text("{}", encoding="utf-8")
    queued = pending / "abcdef123456.json"
    queued.write_text("{}", encoding="utf-8")
    os.chmod(runtime_home, 0o755)
    os.chmod(pending, 0o755)
    os.chmod(config, 0o644)
    os.chmod(queued, 0o644)
    module = load_module(monkeypatch, runtime_home)

    module.ensure_runtime_dirs()
    module.ensure_runtime_config()

    assert stat.S_IMODE(runtime_home.stat().st_mode) == 0o700
    assert stat.S_IMODE(pending.stat().st_mode) == 0o700
    assert stat.S_IMODE(config.stat().st_mode) == 0o600
    assert stat.S_IMODE(queued.stat().st_mode) == 0o600


def test_atomic_sensitive_writes_are_private(monkeypatch, tmp_path):
    module = load_module(monkeypatch, tmp_path / "runtime_home")
    module.ensure_runtime_dirs()
    target = Path(module.DATA_DIR) / "synthetic.json"
    target.write_text("old", encoding="utf-8")
    os.chmod(target, 0o644)

    module.atomic_write_json(str(target), {"value": "SYNTHETIC"})

    assert stat.S_IMODE(target.stat().st_mode) == 0o600


def test_append_expense_row_finds_headers_below_first_row(monkeypatch, tmp_path):
    module = load_module(monkeypatch, tmp_path / "runtime_home")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Expense Log"
    for column, header in enumerate(module.EXPECTED_HEADERS, start=1):
        worksheet.cell(row=2, column=column).value = header
    worksheet.cell(row=3, column=1).value = 1
    worksheet.cell(row=3, column=11).value = "https://example.test/old"
    worksheet.cell(row=3, column=12).value = "old.pdf"

    buffer = io.BytesIO()
    workbook.save(buffer)
    submission = {
        "date": "2030-01-02",
        "vendor": "SYNTHETIC VENDOR",
        "amount": 10.0,
        "category": "SYNTHETIC CATEGORY 1",
        "business_purpose": "SYNTHETIC BUSINESS PURPOSE",
        "paid_by": "SYNTHETIC USER",
        "payment_method": "SYNTHETIC PAYMENT 1",
        "reimbursable": "Yes",
        "reimbursement_status": "SYNTHETIC STATUS 1",
        "notes": "SYNTHETIC NOTE",
    }

    updated, row_number, expense_id = module.append_expense_row(
        buffer.getvalue(),
        "Expense Log",
        submission,
        "https://example.test/new",
        "2030-01-02_SYNTHETIC-VENDOR_10.00_SYNTHETIC-PURPOSE.pdf",
    )

    assert row_number == 4
    assert expense_id == 2
    saved = load_workbook(io.BytesIO(updated), data_only=False)
    saved_worksheet = saved["Expense Log"]
    assert saved_worksheet.cell(row=4, column=1).value == 2
    assert saved_worksheet.cell(row=4, column=3).value == "SYNTHETIC VENDOR"
    assert saved_worksheet.cell(row=4, column=11).value == "https://example.test/new"


def test_tracked_workbook_template_is_synthetic_and_schema_compatible():
    template = SCRIPT_PATH.parent / "docs" / "Expense_Tracker_Template.xlsx"
    workbook = load_workbook(template, data_only=False)

    assert workbook.sheetnames == [
        "Categories",
        "Expense Log",
        "Summary",
        "Entry Form",
        "Guidelines",
    ]
    assert [cell.value for cell in workbook["Expense Log"][1]] == [
        "ID",
        "Date",
        "Vendor",
        "Amount",
        "Category",
        "Business Purpose",
        "Paid By",
        "Payment Method",
        "Reimbursable?",
        "Reimb. Status",
        "Receipt Link",
        "Receipt Filename",
        "Notes",
    ]
    assert all(cell.value is None for cell in workbook["Expense Log"][2])
    for row in workbook["Categories"].iter_rows(min_row=4, max_row=5, min_col=1, max_col=4):
        assert all(str(cell.value).startswith("SYNTHETIC ") for cell in row)
    assert workbook["Summary"]["B4"].value.startswith("=COUNT(")
    assert workbook["Summary"]["B5"].value.startswith("=SUM(")
    assert workbook.properties.creator in (None, "", "openpyxl")

from __future__ import annotations

import csv
import importlib.util
import math
import os
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest


PROJECT_DIR = Path(__file__).resolve().parents[1]
LAUNCHER = PROJECT_DIR / "hysa_vs_cd_model.py"
MODULE_SPEC = importlib.util.spec_from_file_location("hysa_vs_cd_model", LAUNCHER)
assert MODULE_SPEC and MODULE_SPEC.loader
MODEL = importlib.util.module_from_spec(MODULE_SPEC)
MODULE_SPEC.loader.exec_module(MODEL)

VALID_ROWS = [
    ("Initial Principal", "0"),
    ("Starting HYSA Rate", "0"),
    ("Starting CD Rate", "0"),
    ("Rate Step (per period)", "0"),
    # Only structural workbook controls are positive in test fixtures.
    ("Rate Change Frequency (months)", "1"),
    ("CD Sensitivity", "0"),
    ("Total Duration (months)", "2"),
]


def write_inputs(path: Path, rows: list[tuple[str, str]] = VALID_ROWS) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Parameter", "Value"])
        writer.writerows(rows)


def with_values(**overrides: str) -> list[tuple[str, str]]:
    return [(name, overrides.get(name, value)) for name, value in VALID_ROWS]


def run_cli(tmp_path: Path, *args: str) -> subprocess.CompletedProcess[str]:
    runtime_home = tmp_path / "runtime"
    env = os.environ.copy()
    env["HYSA_EXCEL_HOME"] = str(runtime_home)
    return subprocess.run(
        [sys.executable, str(LAUNCHER), *args],
        cwd=tmp_path,
        env=env,
        text=True,
        capture_output=True,
        check=False,
    )


def test_first_run_creates_incomplete_template_and_no_workbook(tmp_path: Path) -> None:
    result = run_cli(tmp_path)

    template = tmp_path / "runtime" / "inputs.csv"
    output = tmp_path / "runtime" / "CD_vs_HYSA_Model.xlsx"
    assert result.returncode == 2
    assert "CONFIGURATION REQUIRED" in result.stderr
    assert str(template) in result.stderr
    assert template.exists()
    assert not output.exists()

    with template.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert [row["Parameter"] for row in rows] == [name for name, _ in VALID_ROWS]
    assert all(row["Value"] == "" for row in rows)


def test_help_does_not_create_runtime_state(tmp_path: Path) -> None:
    result = run_cli(tmp_path, "--help")

    assert result.returncode == 0
    assert "--inputs" in result.stdout
    assert "--output" in result.stdout
    assert not (tmp_path / "runtime").exists()


def test_tracked_example_is_all_zero_and_intentionally_invalid(tmp_path: Path) -> None:
    example = PROJECT_DIR / "inputs.example.csv"
    with example.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert rows
    assert all(row["Value"] == "0" for row in rows)

    result = run_cli(tmp_path, "--inputs", str(example))

    assert result.returncode == 2
    assert "Rate Change Frequency (months): must be greater than zero" in result.stderr
    assert "Total Duration (months): must be greater than zero" in result.stderr
    assert not (tmp_path / "runtime" / "CD_vs_HYSA_Model.xlsx").exists()


def test_missing_explicit_input_creates_template_at_that_path(tmp_path: Path) -> None:
    explicit_inputs = tmp_path / "private" / "scenario.csv"
    explicit_output = tmp_path / "results" / "scenario.xlsx"

    result = run_cli(
        tmp_path,
        "--inputs",
        str(explicit_inputs),
        "--output",
        str(explicit_output),
    )

    assert result.returncode == 2
    assert "CONFIGURATION REQUIRED" in result.stderr
    assert explicit_inputs.exists()
    assert not explicit_output.exists()
    assert not explicit_output.parent.exists()


def test_adjacent_legacy_inputs_are_not_migrated(tmp_path: Path) -> None:
    write_inputs(tmp_path / "inputs.csv")

    result = run_cli(tmp_path)

    runtime_template = tmp_path / "runtime" / "inputs.csv"
    assert result.returncode == 2
    assert runtime_template.exists()
    assert "CONFIGURATION REQUIRED" in result.stderr
    assert not (tmp_path / "runtime" / "CD_vs_HYSA_Model.xlsx").exists()
    with runtime_template.open(newline="", encoding="utf-8") as handle:
        assert all(row["Value"] == "" for row in csv.DictReader(handle))


def test_incomplete_config_has_actionable_error_and_no_output(tmp_path: Path) -> None:
    inputs = tmp_path / "runtime" / "inputs.csv"
    write_inputs(inputs, [(name, "") for name, _ in VALID_ROWS])

    result = run_cli(tmp_path)

    assert result.returncode == 2
    assert "CONFIGURATION REQUIRED" in result.stderr
    assert "Initial Principal" in result.stderr
    assert not (tmp_path / "runtime" / "CD_vs_HYSA_Model.xlsx").exists()


def test_explicit_input_and_output_paths_generate_workbook(tmp_path: Path) -> None:
    inputs = tmp_path / "scenario.csv"
    output = tmp_path / "results" / "comparison.xlsx"
    write_inputs(inputs)

    result = run_cli(tmp_path, "--inputs", str(inputs), "--output", str(output))

    assert result.returncode == 0, result.stderr
    assert output.exists()
    assert str(output) in result.stdout
    assert not (tmp_path / "runtime").exists()


def test_default_output_is_written_to_runtime_home(tmp_path: Path) -> None:
    inputs = tmp_path / "runtime" / "inputs.csv"
    write_inputs(inputs)

    result = run_cli(tmp_path)

    output = tmp_path / "runtime" / "CD_vs_HYSA_Model.xlsx"
    assert result.returncode == 0, result.stderr
    assert output.exists()


def test_validation_rejects_duplicates_and_non_integer_periods(tmp_path: Path) -> None:
    inputs = tmp_path / "invalid.csv"
    rows = VALID_ROWS + [("Initial Principal", "0")]
    rows = [
        (name, "1.5" if name == "Total Duration (months)" else value)
        for name, value in rows
    ]
    write_inputs(inputs, rows)

    result = run_cli(tmp_path, "--inputs", str(inputs))

    assert result.returncode == 2
    assert "duplicate parameter" in result.stderr.lower()
    assert "whole number" in result.stderr.lower()
    assert not (tmp_path / "runtime" / "CD_vs_HYSA_Model.xlsx").exists()


def test_output_cannot_overwrite_the_input_config(tmp_path: Path) -> None:
    inputs = tmp_path / "private-inputs.csv"
    write_inputs(inputs)
    original = inputs.read_bytes()

    result = run_cli(tmp_path, "--inputs", str(inputs), "--output", str(inputs))

    assert result.returncode == 2
    assert "must be different" in result.stderr
    assert inputs.read_bytes() == original


def test_generated_workbook_has_four_formula_driven_sheets(tmp_path: Path) -> None:
    inputs = tmp_path / "scenario.csv"
    output = tmp_path / "comparison.xlsx"
    write_inputs(inputs)
    result = run_cli(tmp_path, "--inputs", str(inputs), "--output", str(output))
    assert result.returncode == 0, result.stderr

    workbook = openpyxl.load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Inputs", "Monthly Balances", "Simple", "Output"]
    assert workbook["Inputs"]["A2"].value == "Initial Principal"
    assert workbook["Inputs"]["B2"].value == 0
    month_formula = workbook["Monthly Balances"]["A2"].value
    month_formula_text = getattr(month_formula, "text", month_formula)
    assert "SEQUENCE(" in month_formula_text
    assert workbook["Monthly Balances"]["B2"].value.startswith("=MAX(")
    assert workbook["Monthly Balances"]["C2"].value == "='Inputs'!$B$2*(1+B2/12)"
    assert workbook["Monthly Balances"]["C3"].value == "=C2*(1+B3/12)"
    assert workbook["Output"]["B2"].value == "='Monthly Balances'!C3"
    assert "MAX($B$2:$B$9)" in workbook["Output"]["C2"].value


def test_one_month_duration_compounds_once_and_output_uses_that_balance(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "one-month.csv"
    output = tmp_path / "one-month.xlsx"
    write_inputs(inputs, with_values(**{"Total Duration (months)": "1"}))

    result = run_cli(tmp_path, "--inputs", str(inputs), "--output", str(output))
    assert result.returncode == 0, result.stderr

    workbook = openpyxl.load_workbook(output, data_only=False)
    balances = workbook["Monthly Balances"]
    assert balances["C2"].value == "='Inputs'!$B$2*(1+B2/12)"
    assert balances["E2"].value == "='Inputs'!$B$2*(1+D2/12)"
    assert workbook["Output"]["B2"].value == "='Monthly Balances'!C2"
    assert workbook["Output"]["B3"].value == "='Monthly Balances'!E2"


def test_multi_month_compounding_has_n_periods_numerically_and_in_formulas(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "multi-month.csv"
    output = tmp_path / "multi-month.xlsx"
    write_inputs(inputs, with_values(**{"Total Duration (months)": "3"}))

    result = run_cli(tmp_path, "--inputs", str(inputs), "--output", str(output))
    assert result.returncode == 0, result.stderr

    workbook = openpyxl.load_workbook(output, data_only=False)
    balances = workbook["Monthly Balances"]
    assert balances["C2"].value == "='Inputs'!$B$2*(1+B2/12)"
    assert balances["C3"].value == "=C2*(1+B3/12)"
    assert balances["C4"].value == "=C3*(1+B4/12)"
    assert workbook["Output"]["B2"].value == "='Monthly Balances'!C4"

    synthetic_principal = 1e100
    synthetic_rate = 1.0
    projected = MODEL.compound_monthly(
        synthetic_principal,
        [synthetic_rate, synthetic_rate, synthetic_rate],
    )
    assert math.isclose(
        projected,
        synthetic_principal * (1 + synthetic_rate / 12) ** 3,
        rel_tol=1e-12,
    )


def test_numeric_safety_bounds_accept_extreme_but_safe_inputs(tmp_path: Path) -> None:
    inputs = tmp_path / "safe-boundary.csv"
    write_inputs(
        inputs,
        with_values(
            **{
                "Initial Principal": "1e100",
                "Starting HYSA Rate": "1",
                "Starting CD Rate": "1",
                "CD Sensitivity": "100",
                "Total Duration (months)": "1200",
            }
        ),
    )

    values = MODEL.load_inputs(inputs)
    bound = MODEL.maximum_compounded_balance(
        values["Initial Principal"],
        1.0,
        int(values["Total Duration (months)"]),
    )

    assert math.isfinite(bound)
    assert bound < MODEL.SAFE_BALANCE_LIMIT


def test_numeric_safety_rejects_principal_and_sensitivity_above_bounds(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "unsafe-static-bounds.csv"
    write_inputs(
        inputs,
        with_values(
            **{
                "Initial Principal": "1.01e100",
                "CD Sensitivity": "100.01",
            }
        ),
    )

    try:
        MODEL.load_inputs(inputs)
    except MODEL.ConfigurationError as exc:
        message = str(exc)
    else:
        raise AssertionError("unsafe static bounds were accepted")

    assert "Initial Principal: must be 1e+100 or less" in message
    assert "CD Sensitivity: must be 100 or less" in message


def test_numeric_safety_rejects_projected_rate_growth(tmp_path: Path) -> None:
    inputs = tmp_path / "unsafe-rate-growth.csv"
    write_inputs(
        inputs,
        with_values(
            **{
                "Rate Step (per period)": "1",
                "Rate Change Frequency (months)": "1",
                "CD Sensitivity": "2",
                "Total Duration (months)": "3",
            }
        ),
    )

    try:
        MODEL.load_inputs(inputs)
    except MODEL.ConfigurationError as exc:
        message = str(exc)
    else:
        raise AssertionError("unsafe projected rates were accepted")

    assert "projected HYSA rate" in message
    assert "projected CD rate" in message
    assert "100%" in message


def test_template_creation_closes_descriptor_when_setup_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    descriptors: list[int] = []
    original_mkstemp = MODEL.tempfile.mkstemp

    def tracking_mkstemp(*args: object, **kwargs: object) -> tuple[int, str]:
        descriptor, name = original_mkstemp(*args, **kwargs)
        descriptors.append(descriptor)
        return descriptor, name

    def fail_fchmod(descriptor: int, mode: int) -> None:
        raise OSError("synthetic permission failure")

    monkeypatch.setattr(MODEL.tempfile, "mkstemp", tracking_mkstemp)
    monkeypatch.setattr(MODEL.os, "fchmod", fail_fchmod)

    with pytest.raises(OSError, match="synthetic permission failure"):
        MODEL.write_incomplete_template(tmp_path / "inputs.csv")

    assert len(descriptors) == 1
    with pytest.raises(OSError):
        os.fstat(descriptors[0])


def test_workbook_is_closed_when_build_fails_midway(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    inputs = tmp_path / "inputs.csv"
    write_inputs(inputs)
    values = MODEL.load_inputs(inputs)
    add_calls = 0
    closed: list[bool] = []
    original_add = MODEL.xlsxwriter.Workbook.add_worksheet
    original_close = MODEL.xlsxwriter.Workbook.close

    def fail_second_sheet(workbook: object, *args: object, **kwargs: object) -> object:
        nonlocal add_calls
        add_calls += 1
        if add_calls == 2:
            raise RuntimeError("synthetic mid-build failure")
        return original_add(workbook, *args, **kwargs)

    def tracking_close(workbook: object) -> None:
        closed.append(True)
        original_close(workbook)

    monkeypatch.setattr(MODEL.xlsxwriter.Workbook, "add_worksheet", fail_second_sheet)
    monkeypatch.setattr(MODEL.xlsxwriter.Workbook, "close", tracking_close)

    with pytest.raises(RuntimeError, match="synthetic mid-build failure"):
        MODEL.build_workbook(values, tmp_path / "failed.xlsx")

    assert closed == [True]
    assert not (tmp_path / "failed.xlsx").exists()


def test_percent_strings_are_accepted_without_embedding_defaults(tmp_path: Path) -> None:
    inputs = tmp_path / "percent.csv"
    rows = [
        (
            name,
            "0%"
            if name
            in {
                "Starting HYSA Rate",
                "Starting CD Rate",
                "Rate Step (per period)",
                "CD Sensitivity",
            }
            else value,
        )
        for name, value in VALID_ROWS
    ]
    write_inputs(inputs, rows)
    output = tmp_path / "percent.xlsx"

    result = run_cli(tmp_path, "--inputs", str(inputs), "--output", str(output))

    assert result.returncode == 0, result.stderr
    workbook = openpyxl.load_workbook(output, data_only=False)
    assert workbook["Inputs"]["B3"].value == 0
